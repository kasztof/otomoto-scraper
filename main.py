import os
from collections import defaultdict
from urllib.request import urlopen
import statistics

import collections

import xlsxwriter as xlsxwriter
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.chart import BarChart

MAXIMUM_MILEAGE = 2000000  # prevent trolls with mileages like from earth to mars


def get_car_url(mark, model):
    return 'https://www.otomoto.pl/osobowe/' + mark + '/' + model + '/?page='


def merge_dictionaries(dict1, dict2):
    """Function that merges two dictionaries into one dict.
    Keys can be duplicated"""

    result = defaultdict(list)

    for key, value in dict1.items():
        if key in dict2:
            result[key] = [*value, *dict2[key]]
        else:
            result[key] = value

    for key, value in dict2.items():
        if key not in dict1:
            result[key] = value

    return result


def merge_list_of_dictionaries(dict_list):
    """Function that merges list of dictionaries and return one dict with all pairs of k:v"""
    for i in range(1, len(dict_list)):
        dict_list[0] = merge_dictionaries(dict_list[0], dict_list[i])
    return dict_list[0]


def get_max_page_number(url):
    page = urlopen(url + '1')  # we are adding '1' because the url pattern ends on /?page=
    soup = BeautifulSoup(page, "lxml")
    pages = soup.find_all(attrs={"class": "page"})
    max_page = pages[-2].string  # [-2] because the last element with class 'page' is Next page
    #  so we have to pick the one before
    return max_page


def get_mileages_and_years(url):
    """Returns dict with year:mileage pairs from given url"""
    page = urlopen(url)
    soup = BeautifulSoup(page, "lxml")
    mileages = soup.find_all(attrs={"data-code": "mileage"})
    years = soup.find_all(attrs={"data-code": "year"})

    result_dict = defaultdict(list)

    for yrs, mlg in list(zip(years, mileages)):
        mileage = get_number(mlg.span.string)
        if mileage < MAXIMUM_MILEAGE:
            result_dict[yrs.span.string].append(mileage)

    return result_dict


def get_number(string):
    """Returns only the number from pattern: 'mileage km' where mileage is a number"""
    result = ""
    for char in string:
        if char.isdigit():
            result += char
    return int(result)


def dict_with_median_and_average_values(dictionary):
    """Takes dictionary with pairs of year:[list of mileages]
     and return a dict with pairs year:[median of mileages, avg of mileages]"""
    result = defaultdict(list)
    for key, val in dictionary.items():
        median = statistics.median(val)
        avg = statistics.mean(val)
        result[key].append(median)
        result[key].append(avg)

    return result


def init_xlsx(path):
    if os.path.exists(path + '/data.xlsx'):
        workbook = openpyxl.load_workbook(path + '/data.xlsx')
    else:
        wb = xlsxwriter.Workbook(path + '/data.xlsx')  # if xlsx doesnt exits create it and open with openpyxl
        wb.close()
        workbook = openpyxl.load_workbook(path + '/data.xlsx')
    return workbook


def init_car_sheet(workbook, mark, model, sum_of_pages, data_to_save):
    """Fills sheet with given data"""
    if (mark + '_' + model) not in workbook.sheetnames:
        workbook.create_sheet(mark + '_' + model)
        sheet = workbook[mark + '_' + model]
    else:
        sheet = workbook[mark + '_' + model]

    sheet['D1'] = 'data gathered from: ' + str(sum_of_pages) + ' pages'
    sheet['B1'] = 'median'
    sheet['C1'] = 'average'

    for key, value in data_to_save.items():
        sheet.append([key, *value])

    return sheet


def add_charts(sheet, mark, model, data, cats):
    """draw charts with median and average mileages for each year"""
    chart = BarChart()
    chart.title = mark + " " + model
    chart.type = "col"
    chart.y_axis.title = 'Mileage'
    chart.x_axis.title = 'Year of prod.'

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    sheet.add_chart(chart, "E3")


def save_to_xlsx(path, data, mark, model, sum_of_pages):
    data_to_save = collections.OrderedDict(
        sorted(dict_with_median_and_average_values(data).items()))  # sort dict by years

    workbook = init_xlsx(path)
    sheet = init_car_sheet(workbook, mark, model, sum_of_pages, data_to_save)

    data = openpyxl.chart.Reference(sheet, min_col=2, max_col=3, min_row=1, max_row=len(data_to_save))
    cats = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_row=len(data_to_save))

    add_charts(sheet, mark, model, data, cats)

    workbook.save(path + '/data.xlsx')
